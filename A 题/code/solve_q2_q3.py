# -*- coding: utf-8 -*-
"""A 题问题 2、3：负载均衡与分时圆形禁飞区下的多无人机调度。

问题 2 固定问题 1 已采用的机队规模，采用分层优化：第一层以问题 1 的
最短总完成时间 Tmax* 为基准；第二层在 Tmax <= Tmax*(1+eps) 容差上限内
最小化工作负载极差 delta=Tmax-Tmin，并显式使用“最忙路径搬点到最闲路径”
的均衡算子。问题 3 仍固定该机队规模，从 08:00 同时开始执行。对每条航段
进行连续时间事件仿真，并在“入界前等待”和“沿圆外切线—安全圆弧绕飞”
之间选较早到达者。

算法是多起点模拟退火/变邻域搜索，任务副本始终只移动、不增删，因此覆盖次数
保持不变；相邻同一物理点的方案被判为非法。输出 result2.xlsx、result3.xlsx，
并生成 docs/q2_q3_summary.txt 供独立核验。

运行：python solve_q2_q3.py
依赖：openpyxl（以及 solve_q1.py 的数据读取常量）
"""

from __future__ import annotations

import math
import os
import random
from collections import Counter
from dataclasses import dataclass
from datetime import time as dt_time
from functools import lru_cache
from typing import Callable, Iterable, Sequence

import openpyxl

from solve_q1 import (
    BASE_DIR,
    LEVEL_TIMES,
    SCALE_M,
    SERVICE_S,
    SPEED_MS,
    load_case,
)


RESULT1 = os.path.join(BASE_DIR, "result1.xlsx")
RESULT2 = os.path.join(BASE_DIR, "result2.xlsx")
RESULT3 = os.path.join(BASE_DIR, "result3.xlsx")
DATA2 = os.path.join(BASE_DIR, "附件2.xlsx")
SUMMARY = os.path.join(BASE_DIR, "docs", "q2_q3_summary.txt")
EPS = 1e-7
INF = 1e100
RELAX_TOL = 0.01  # 问题 2 第二层的 Tmax 容差: 上限 = Tmax* * (1 + RELAX_TOL)


@dataclass(frozen=True)
class Zone:
    zid: str
    x: float
    y: float
    radius: float
    start: float  # 相对 08:00 的秒数
    end: float


@dataclass(frozen=True)
class RouteMetric:
    total: float
    flight: float
    service: float
    wait: float


def clock_seconds(value) -> float:
    """把 Excel 中的 HH:MM / datetime.time 转成相对 08:00 的秒数。"""
    if isinstance(value, dt_time):
        hour, minute, second = value.hour, value.minute, value.second
    else:
        parts = str(value).strip().split(":")
        hour, minute = int(parts[0]), int(parts[1])
        second = int(float(parts[2])) if len(parts) > 2 else 0
    return (hour - 8) * 3600.0 + minute * 60.0 + second


def load_zones(sheet: str) -> list[Zone]:
    wb = openpyxl.load_workbook(DATA2, data_only=True, read_only=True)
    ws = wb[sheet]
    zones = []
    for zid, x, y, radius, start, end in ws.iter_rows(min_row=2, values_only=True):
        if zid is None:
            continue
        st, en = clock_seconds(start), clock_seconds(end)
        # 附件 Case4 的 Z8 为 17:00--17:00，零长度窗口不构成管制时段。
        if en > st + EPS:
            zones.append(Zone(str(zid), float(x), float(y), float(radius), st, en))
    wb.close()
    return zones


def load_routes(path: str, sheet: str) -> list[list[int]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    routes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            routes.append([int(v) for v in row[1:] if v is not None])
    wb.close()
    return routes


def valid_route(route: Sequence[int]) -> bool:
    return bool(route) and all(a != b for a, b in zip(route, route[1:]))


@lru_cache(maxsize=500_000)
def segment_disc_interval(a: tuple[float, float], b: tuple[float, float],
                          z: Zone) -> tuple[float, float] | None:
    """线段 a->b 位于闭圆内的参数区间 [u,v]，参数范围为 [0,1]。"""
    ax, ay = a
    dx, dy = b[0] - ax, b[1] - ay
    fx, fy = ax - z.x, ay - z.y
    aa = dx * dx + dy * dy
    cc = fx * fx + fy * fy - z.radius * z.radius
    if aa <= EPS:
        return (0.0, 1.0) if cc <= EPS else None
    bb = 2.0 * (fx * dx + fy * dy)
    disc = bb * bb - 4.0 * aa * cc
    if disc < -EPS:
        return None
    root = math.sqrt(max(0.0, disc))
    lo = max(0.0, (-bb - root) / (2.0 * aa))
    hi = min(1.0, (-bb + root) / (2.0 * aa))
    return (lo, hi) if lo <= hi + EPS else None


def straight_timed_leg(start_time: float, a: tuple[float, float], b: tuple[float, float],
                       destination_service: float, zones: Sequence[Zone],
                       destination_clearance: dict[str, float] | None = None) -> RouteMetric:
    """沿直线航段飞行；必要时在圆区入界前等待，返回到达并完成服务后的指标。

    对每个圆，线段在圆内的飞行时间是连续区间。若该区间（终点在圆内时连同
    终点服务）与管制窗口有交集，就把等待安排在其入界点外。按入界位置排序后
    逐个处理，已增加的等待会自动推迟后续圆区的到达时刻。
    """
    distance_m = math.dist(a, b) * SCALE_M
    duration = distance_m / SPEED_MS
    events = []
    clearance = destination_clearance or {}
    for z in zones:
        interval = segment_disc_interval(a, b, z)
        if interval is None:
            continue
        u, v = interval
        inside_tail = (
            destination_service + clearance.get(z.zid, 0.0)
            if v >= 1.0 - EPS else 0.0
        )
        events.append((u, v, inside_tail, z))
    events.sort(key=lambda item: (item[0], item[1], item[3].start))

    waiting = 0.0
    for u, v, inside_tail, z in events:
        enter = start_time + waiting + u * duration
        leave = start_time + waiting + v * duration + inside_tail
        # 接触边界按禁飞处理；等待至 end+EPS 后进入。
        if enter < z.end - EPS and leave > z.start + EPS:
            # 航段起点已在圆内时，不能把等待安排在圆内；这种时序应由前一
            # 入界航段的“服务+下一段离界前视”提前推迟到解禁后。
            if u <= EPS:
                return RouteMetric(INF, INF, INF, INF)
            # 等待点取入界点外侧：沿航段从入界点后退，直到该点不落入任何
            # 生效禁飞圆内部（重叠圆区需要在共同外侧等待）。
            u_wait = u
            seg_len = max(math.dist(a, b), 1e-9)
            for _ in range(10000):
                pw = (a[0] + (b[0] - a[0]) * u_wait, a[1] + (b[1] - a[1]) * u_wait)
                wait_start = start_time + waiting + u_wait * duration
                safe = True
                for z2 in zones:
                    if (wait_start < z2.end - EPS and z.end > z2.start + EPS
                            and math.dist(pw, (z2.x, z2.y)) < z2.radius - EPS):
                        safe = False
                        break
                if safe:
                    break
                u_wait -= 1.0 / seg_len  # 每次后退 1 坐标单位 (100 m)
                if u_wait <= EPS:
                    return RouteMetric(INF, INF, INF, INF)
            waiting += z.end - (start_time + waiting + u_wait * duration) + EPS

    return RouteMetric(
        total=duration + destination_service + waiting,
        flight=duration,
        service=destination_service,
        wait=waiting,
    )


@lru_cache(maxsize=200_000)
def tangent_detours(a: tuple[float, float], b: tuple[float, float],
                    z: Zone) -> tuple[tuple[tuple[float, float], ...], ...]:
    """给出绕单个圆的两条较短切线—离散圆弧候选折线（不含起终点）。"""
    # 1 个坐标单位即 100m；外扩 1 单位可抵消圆弧离散弦的内缩误差。
    radius = z.radius + 1.0
    ca = (a[0] - z.x, a[1] - z.y)
    cb = (b[0] - z.x, b[1] - z.y)
    da, db = math.hypot(*ca), math.hypot(*cb)
    if da <= radius + EPS or db <= radius + EPS:
        return ()
    alpha, beta = math.atan2(ca[1], ca[0]), math.atan2(cb[1], cb[0])
    ta = [alpha - math.acos(radius / da), alpha + math.acos(radius / da)]
    tb = [beta - math.acos(radius / db), beta + math.acos(radius / db)]
    candidates: list[tuple[float, list[tuple[float, float]]]] = []
    max_step = math.pi / 18.0  # 10°，在外扩圆上取弦仍位于原禁飞圆外
    for aa in ta:
        for bb in tb:
            for direction in (1, -1):
                delta = (bb - aa) % (2.0 * math.pi)
                if direction < 0:
                    delta = -((aa - bb) % (2.0 * math.pi))
                steps = max(1, int(math.ceil(abs(delta) / max_step)))
                angles = [aa + delta * k / steps for k in range(steps + 1)]
                path = [(z.x + radius * math.cos(t), z.y + radius * math.sin(t)) for t in angles]
                length = math.dist(a, path[0]) + sum(
                    math.dist(p, q) for p, q in zip(path, path[1:])
                ) + math.dist(path[-1], b)
                candidates.append((length, path))
    candidates.sort(key=lambda item: item[0])
    # 两侧绕行通常对应最短的两个候选；去除数值上重复的路径。
    answer: list[tuple[tuple[float, float], ...]] = []
    signatures = set()
    for _, path in candidates:
        sig = tuple((round(x, 5), round(y, 5)) for x, y in path)
        if sig not in signatures:
            signatures.add(sig)
            answer.append(tuple(path))
        if len(answer) == 2:
            break
    return tuple(answer)


def polyline_metric(start_time: float, points: Sequence[tuple[float, float]],
                    destination_service: float, zones: Sequence[Zone],
                    destination_clearance: dict[str, float] | None = None) -> RouteMetric:
    """用直线事件评价器逐段评价给定折线，服务只在最后一个点发生。"""
    now = start_time
    flight = service = waiting = 0.0
    for i, (a, b) in enumerate(zip(points, points[1:])):
        serv = destination_service if i == len(points) - 2 else 0.0
        clearance = destination_clearance if serv > 0.0 else None
        m = straight_timed_leg(now, a, b, serv, zones, clearance)
        if m.total >= INF / 2:
            return RouteMetric(INF, INF, INF, INF)
        now += m.total
        flight += m.flight
        service += m.service
        waiting += m.wait
    return RouteMetric(now - start_time, flight, service, waiting)


def timed_leg(start_time: float, a: tuple[float, float], b: tuple[float, float],
              destination_service: float, zones: Sequence[Zone],
              destination_clearance: dict[str, float] | None = None) -> RouteMetric:
    """返回等待直飞与切线绕飞候选中的最早到达方案。"""
    best = straight_timed_leg(
        start_time, a, b, destination_service, zones, destination_clearance
    )
    if best.wait <= EPS:
        return best
    for z in zones:
        if segment_disc_interval(a, b, z) is None:
            continue
        for detour in tangent_detours(a, b, z):
            candidate = polyline_metric(
                start_time, [a, *detour, b], destination_service, zones,
                destination_clearance,
            )
            if candidate.total < best.total - EPS:
                best = candidate
    return best


def static_metric(route: Sequence[int], coords: dict[int, tuple[float, float]]) -> RouteMetric:
    if not valid_route(route):
        return RouteMetric(INF, INF, INF, INF)
    flight = 0.0
    previous = (0.0, 0.0)
    for pid in route:
        point = coords[pid]
        flight += math.dist(previous, point) * SCALE_M / SPEED_MS
        previous = point
    flight += math.dist(previous, (0.0, 0.0)) * SCALE_M / SPEED_MS
    service = len(route) * SERVICE_S
    return RouteMetric(flight + service, flight, service, 0.0)


def clearance_after_visit(route: Sequence[int], position: int,
                          coords: dict[int, tuple[float, float]], z: Zone) -> float:
    """抵达 route[position] 并完成本次服务后，沿后续路线首次离开 z 的时间。

    当前点的服务时间由入界航段单独计入；返回值包括后续圆内飞行，以及若后续
    巡检点仍在圆内时的服务时间。该前视用于把整段连续圆内占用统一安排在管制
    窗口之前或之后，避免在圆内任务链中途非法等待。
    """
    clearance = 0.0
    a = coords[route[position]]
    for next_position in range(position + 1, len(route) + 1):
        is_base = next_position == len(route)
        b = (0.0, 0.0) if is_base else coords[route[next_position]]
        duration = math.dist(a, b) * SCALE_M / SPEED_MS
        interval = segment_disc_interval(a, b, z)
        if interval is None or interval[0] > EPS:
            return clearance
        clearance += interval[1] * duration
        if interval[1] < 1.0 - EPS or is_base:
            return clearance
        clearance += SERVICE_S
        a = b
    return clearance


def dynamic_metric(route: Sequence[int], coords: dict[int, tuple[float, float]],
                   zones: Sequence[Zone]) -> RouteMetric:
    if not valid_route(route):
        return RouteMetric(INF, INF, INF, INF)
    now = flight = service = waiting = 0.0
    previous = (0.0, 0.0)
    for position, pid in enumerate(route):
        destination = coords[pid]
        clearance: dict[str, float] = {}
        for z in zones:
            value = clearance_after_visit(route, position, coords, z)
            if value > EPS:
                clearance[z.zid] = value
        m = timed_leg(now, previous, destination, SERVICE_S, zones, clearance)
        if m.total >= INF / 2:
            return RouteMetric(INF, INF, INF, INF)
        now += m.total
        flight += m.flight
        service += m.service
        waiting += m.wait
        previous = destination
    m = timed_leg(now, previous, (0.0, 0.0), 0.0, zones)
    if m.total >= INF / 2:
        return RouteMetric(INF, INF, INF, INF)
    return RouteMetric(now + m.total, flight + m.flight, service, waiting + m.wait)


def objective(times: Sequence[float], balance_weight: float) -> float:
    """主目标 Tmax；次目标极差以较小权重参与搜索。"""
    if not times or max(times) >= INF / 2:
        return INF
    return max(times) + balance_weight * (max(times) - min(times))


def hierarchical_objective(times: Sequence[float], tmax_cap: float,
                           relax: float = 0.0) -> float:
    """分层目标（问题 2 第二层）：
    第一层：Tmax 不得超过上限 cap = tmax_cap * (1 + relax)，越界部分乘大罚系数；
    第二层：可行域内最小化工作负载极差 delta = Tmax - Tmin。
    """
    if not times or max(times) >= INF / 2:
        return INF
    mx, mn = max(times), min(times)
    cap = tmax_cap * (1.0 + relax)
    over = max(0.0, mx - cap)
    return over * 1e6 + (mx - mn)


def lex_key(times: Sequence[float]) -> tuple[float, float, float]:
    """保存最优解时采用严格词典序，再以方差作第三判据。"""
    mx, mn = max(times), min(times)
    mean = sum(times) / len(times)
    return mx, mx - mn, sum((x - mean) ** 2 for x in times)


def best_cyclic_route(route: Sequence[int],
                      metric: Callable[[Sequence[int]], RouteMetric]) -> tuple[list[int], RouteMetric]:
    """穷举正反向循环切入位置，返回该任务环的最佳起止点。"""
    best_route = list(route)
    best_metric = metric(best_route)
    for oriented in (list(route), list(reversed(route))):
        for cut in range(len(oriented)):
            candidate = oriented[cut:] + oriented[:cut]
            if not valid_route(candidate):
                continue
            candidate_metric = metric(candidate)
            if candidate_metric.total < best_metric.total - EPS:
                best_route, best_metric = candidate, candidate_metric
    return best_route, best_metric


def cyclic_improve(routes: Sequence[Sequence[int]],
                   metric: Callable[[Sequence[int]], RouteMetric]) -> tuple[list[list[int]], list[RouteMetric]]:
    improved = [best_cyclic_route(route, metric) for route in routes]
    return [x[0] for x in improved], [x[1] for x in improved]


def mutate(routes: list[list[int]], rng: random.Random) -> tuple[list[list[int]], set[int]]:
    """生成一个保持任务副本总集合不变的邻域解。"""
    candidate = [r[:] for r in routes]
    n = len(candidate)
    move = rng.random()
    changed: set[int]

    if move < 0.18:  # 路径内 2-opt
        i = rng.randrange(n)
        if len(candidate[i]) < 3:
            return candidate, {i}
        a, b = sorted(rng.sample(range(len(candidate[i])), 2))
        candidate[i][a:b + 1] = reversed(candidate[i][a:b + 1])
        changed = {i}
    elif move < 0.31:  # 路径内 relocate
        i = rng.randrange(n)
        if len(candidate[i]) < 2:
            return candidate, {i}
        a = rng.randrange(len(candidate[i]))
        value = candidate[i].pop(a)
        b = rng.randrange(len(candidate[i]) + 1)
        candidate[i].insert(b, value)
        changed = {i}
    elif move < 0.49:  # 跨路径 relocate
        i, j = rng.sample(range(n), 2)
        if len(candidate[i]) <= 1:
            return candidate, {i, j}
        a = rng.randrange(len(candidate[i]))
        value = candidate[i].pop(a)
        b = rng.randrange(len(candidate[j]) + 1)
        candidate[j].insert(b, value)
        changed = {i, j}
    elif move < 0.64:  # 路径内或跨路径 swap
        i, j = rng.randrange(n), rng.randrange(n)
        if not candidate[i] or not candidate[j]:
            return candidate, {i, j}
        a, b = rng.randrange(len(candidate[i])), rng.randrange(len(candidate[j]))
        candidate[i][a], candidate[j][b] = candidate[j][b], candidate[i][a]
        changed = {i, j}
    elif move < 0.78:  # 成段跨路径 relocate，跨越时间窗目标的台阶
        i, j = rng.sample(range(n), 2)
        if len(candidate[i]) <= 2:
            return candidate, {i, j}
        length = rng.randint(2, min(10, len(candidate[i]) - 1))
        a = rng.randrange(len(candidate[i]) - length + 1)
        block = candidate[i][a:a + length]
        del candidate[i][a:a + length]
        b = rng.randrange(len(candidate[j]) + 1)
        if rng.random() < 0.5:
            block.reverse()
        candidate[j][b:b] = block
        changed = {i, j}
    elif move < 0.90:  # 两条路径交换短任务块
        i, j = rng.sample(range(n), 2)
        li = rng.randint(1, min(6, len(candidate[i])))
        lj = rng.randint(1, min(6, len(candidate[j])))
        a = rng.randrange(len(candidate[i]) - li + 1)
        b = rng.randrange(len(candidate[j]) - lj + 1)
        bi, bj = candidate[i][a:a + li], candidate[j][b:b + lj]
        candidate[i][a:a + li], candidate[j][b:b + lj] = bj, bi
        changed = {i, j}
    else:  # 改变任务环相对基地的切入点，可整体平移到达时序
        i = rng.randrange(n)
        cut = rng.randrange(len(candidate[i]))
        candidate[i] = candidate[i][cut:] + candidate[i][:cut]
        if rng.random() < 0.35:
            candidate[i].reverse()
        changed = {i}
    return candidate, changed


def rebalance_relocate(routes: list[list[int]],
                       metric: Callable[[Sequence[int]], RouteMetric]
                       ) -> tuple[list[list[int]], set[int]]:
    """显式均衡算子：从最忙路径搬一个点插入最闲路径的最佳位置。
    以 (Tmax, delta) 词典序接受改进（不增 Tmax 且极差下降）。
    """
    times = [metric(r).total for r in routes]
    jmax = int(max(range(len(routes)), key=lambda i: times[i]))
    jmin = int(min(range(len(routes)), key=lambda i: times[i]))
    if jmax == jmin or len(routes[jmax]) <= 1:
        return routes, {jmax, jmin}
    cand = [r[:] for r in routes]
    best_key = lex_key(times)
    best_move: tuple[int, int] | None = None
    for i in range(len(cand[jmax])):
        value = cand[jmax][i]
        for pos in range(len(cand[jmin]) + 1):
            trial = [r[:] for r in cand]
            trial[jmax].pop(i)
            trial[jmin].insert(pos, value)
            if not valid_route(trial[jmax]) or not valid_route(trial[jmin]):
                continue
            t_times = times[:]
            t_times[jmax] = metric(trial[jmax]).total
            t_times[jmin] = metric(trial[jmin]).total
            key = lex_key(t_times)
            if key < best_key:
                best_key = key
                best_move = (i, pos)
    if best_move is None:
        return routes, {jmax, jmin}
    i, pos = best_move
    value = cand[jmax].pop(i)
    cand[jmin].insert(pos, value)
    return cand, {jmax, jmin}


def _fix_adjacent(route: Sequence[int], rng: random.Random) -> list[int]:
    """消除路径中相邻的重复点（连续停留只算一次巡检 -> 相邻重复非法）。"""
    out: list[int] = []
    pending: list[int] = []
    last = None
    for pid in route:
        if pid != last:
            out.append(pid)
            last = pid
        else:
            pending.append(pid)
    for pid in pending:
        placed = False
        for attempt in range(len(out) + 1):
            pos = (attempt + rng.randrange(len(out) + 1)) % (len(out) + 1)
            left = out[pos - 1] if pos > 0 else None
            right = out[pos] if pos < len(out) else None
            if left != pid and right != pid:
                out.insert(pos, pid)
                placed = True
                break
        if not placed:
            out.append(pid)
    return out


def _crossover(parent_a: Sequence[Sequence[int]], parent_b: Sequence[Sequence[int]],
               rng: random.Random) -> list[list[int]]:
    """路径继承交叉：子代随机继承父代 A 的若干条路径，其余任务副本取自
    父代 B（按预算过滤后）重新分配到空路径；保持任务副本总集合不变。"""
    n = len(parent_a)
    used: Counter[int] = Counter()
    child: list[list[int] | None] = [None] * n
    for k in range(n):
        if rng.random() < 0.5:
            child[k] = list(parent_a[k])
            for pid in child[k]:
                used[pid] += 1
    pool = [pid for route in parent_b for pid in route]
    need = Counter(pool)
    for pid, count in used.items():
        need[pid] -= count
    take: list[int] = []
    for pid in pool:
        if need[pid] > 0:
            take.append(pid)
            need[pid] -= 1
    empty = [k for k in range(n) if child[k] is None]
    if empty:
        if len(take) < len(empty):
            # 继承过多导致剩余副本不足：放弃继承，全部重新分配
            child = [None] * n
            take = pool[:]
            empty = list(range(n))
        m = len(empty)
        cuts = sorted(rng.sample(range(1, len(take)), m - 1)) if m > 1 else []
        segments: list[list[int]] = []
        prev = 0
        for cut in cuts + [len(take)]:
            segments.append(take[prev:cut])
            prev = cut
        for k, seg in zip(empty, segments):
            child[k] = _fix_adjacent(seg, rng)
    return [r for r in child if r is not None]


def _run_ga(initial: Sequence[Sequence[int]],
            metric: Callable[[Sequence[int]], RouteMetric],
            score_fn: Callable[[Sequence[float]], float],
            *, seed: int, pop_size: int, generations: int,
            tmax_cap: float | None) -> tuple[list[list[int]], list[RouteMetric]]:
    """单次遗传算法运行：种群 + 锦标赛选择 + 路径继承交叉 + 变邻域变异 + 精英保留。"""
    rng = random.Random(seed)
    n = len(initial)
    base = [list(r) for r in initial]

    def evaluate(ind: Sequence[Sequence[int]]) -> tuple[float, list[float]]:
        times = [metric(r).total for r in ind]
        return score_fn(times), times

    # 初始种群：1 个原方案 + 扰动个体
    population = [base]
    for _ in range(pop_size - 1):
        ind = [r[:] for r in base]
        for _ in range(4 + rng.randrange(8)):
            trial, changed = mutate(ind, rng)
            if all(valid_route(trial[k]) for k in changed):
                ind = trial
        population.append(ind)

    fits: list[float] = []
    all_times: list[list[float]] = []
    for ind in population:
        sc, times = evaluate(ind)
        fits.append(sc)
        all_times.append(times)

    best_key = min(lex_key(t) for t in all_times)
    best_i = min(range(pop_size), key=lambda i: lex_key(all_times[i]))
    best_routes = [r[:] for r in population[best_i]]

    elite_n = max(2, pop_size // 10)
    for gen in range(generations):
        order = sorted(range(pop_size), key=lambda i: fits[i])
        new_pop = [[r[:] for r in population[i]] for i in order[:elite_n]]
        while len(new_pop) < pop_size:
            cands = [rng.randrange(pop_size) for _ in range(3)]
            p1 = population[min(cands, key=lambda i: fits[i])]
            cands = [rng.randrange(pop_size) for _ in range(3)]
            p2 = population[min(cands, key=lambda i: fits[i])]
            if rng.random() < 0.65:
                child = _crossover(p1, p2, rng)
            else:
                child = [r[:] for r in p1]
                child, _ = mutate(child, rng)
                child = [_fix_adjacent(r, rng) for r in child]
            new_pop.append(child)
        # 分层模式：每 10 代对最优个体做一次显式均衡搬迁
        if tmax_cap is not None and gen % 10 == 9:
            bi = min(range(pop_size), key=lambda i: fits[i])
            trial, changed = rebalance_relocate(new_pop[bi], metric)
            if all(valid_route(trial[k]) for k in changed):
                new_pop[bi] = trial
        population = new_pop
        fits, all_times = [], []
        for ind in population:
            sc, times = evaluate(ind)
            fits.append(sc)
            all_times.append(times)
        for i, (sc, times) in enumerate(zip(fits, all_times)):
            if sc < INF / 2:
                key = lex_key(times)
                if key < best_key:
                    best_key = key
                    best_routes = [r[:] for r in population[i]]
    return best_routes, [metric(r) for r in best_routes]


def optimise(initial: Sequence[Sequence[int]], metric: Callable[[Sequence[int]], RouteMetric],
             *, seed: int, iterations: int, restarts: int,
             balance_weight: float,
             tmax_cap: float | None = None, relax: float = 0.0
             ) -> tuple[list[list[int]], list[RouteMetric]]:
    """遗传算法 + 变邻域变异，多起点重启取词典序最优。

    若给定 tmax_cap，则采用分层目标（上限内最小化极差），并周期性执行
    “最忙搬点给最闲”的显式均衡算子；否则退化为加权目标 objective。
    iterations 折算为进化代数：种群 48，代数 = max(60, iterations // 120)。
    """
    def score_fn(times: Sequence[float]) -> float:
        if tmax_cap is None:
            return objective(times, balance_weight)
        return hierarchical_objective(times, tmax_cap, relax)

    pop_size = 48
    generations = max(60, iterations // 120)
    global_routes: list[list[int]] | None = None
    global_key: tuple[float, float, float] | None = None
    for restart in range(restarts):
        routes, metrics = _run_ga(
            initial, metric, score_fn,
            seed=seed * 1009 + restart * 9176,
            pop_size=pop_size, generations=generations,
            tmax_cap=tmax_cap,
        )
        key = lex_key([m.total for m in metrics])
        if global_key is None or key < global_key:
            global_key = key
            global_routes = [r[:] for r in routes]
        print(
            f"    restart {restart + 1}/{restarts}: "
            f"Tmax={global_key[0] / 3600:.4f}h, spread={global_key[1] / 3600:.4f}h"
        )
    assert global_routes is not None
    return global_routes, [metric(r) for r in global_routes]


def write_workbook(path: str, results: dict[str, list[list[int]]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet, routes in results.items():
        ws = wb.create_sheet(sheet)
        width = max(len(r) for r in routes)
        ws.append(["UAV ID"] + [f"{i}th Inspection Point" for i in range(1, width + 1)])
        for uid, route in enumerate(routes, 1):
            ws.append([uid] + route)
    wb.save(path)


def cover_signature(routes: Iterable[Sequence[int]]) -> dict[int, int]:
    counts: dict[int, int] = {}
    for route in routes:
        for pid in route:
            counts[pid] = counts.get(pid, 0) + 1
    return counts


def expected_signature(sheet: str) -> dict[int, int]:
    points, _ = load_case(sheet)
    return {int(pid): LEVEL_TIMES[level] for pid, _, _, level in points}


def summary_line(sheet: str, routes: Sequence[Sequence[int]], metrics: Sequence[RouteMetric]) -> str:
    times = [m.total for m in metrics]
    waits = [m.wait for m in metrics]
    assert cover_signature(routes) == expected_signature(sheet)
    assert all(valid_route(r) for r in routes)
    return (
        f"{sheet}: N={len(routes)}, Tmax={max(times)/3600:.6f} h, "
        f"Tmin={min(times)/3600:.6f} h, delta={(max(times)-min(times))/3600:.6f} h, "
        f"total={sum(times)/3600:.6f} h, wait={sum(waits)/3600:.6f} h\n"
        f"  route_hours={[round(x/3600, 6) for x in times]}\n"
        f"  wait_hours={[round(x/3600, 6) for x in waits]}\n"
    )


def main() -> None:
    q2_routes: dict[str, list[list[int]]] = {}
    q2_metrics: dict[str, list[RouteMetric]] = {}
    q3_routes: dict[str, list[list[int]]] = {}
    q3_metrics: dict[str, list[RouteMetric]] = {}

    for idx, sheet in enumerate(("Case1", "Case2", "Case3", "Case4"), 1):
        points, _ = load_case(sheet)
        coords = {int(pid): (float(x), float(y)) for pid, x, y, _ in points}
        initial = load_routes(RESULT1, sheet)
        initial = load_routes(RESULT1, sheet)
        print(f"[{sheet}] 问题2：分层优化（第一层 Tmax*，第二层上限内最小化 delta）")
        metric2 = lambda route, c=coords: static_metric(route, c)
        # 第一层：问题 1 已最小化总完成时间，Tmax* 即问题 1 方案的最长单机时长
        tmax_star = max(metric2(r).total for r in initial)
        # 第二层：在 Tmax <= Tmax*(1+RELAX_TOL) 容差内最小化 delta
        r2, m2 = optimise(
            initial, metric2, seed=20260816 + idx, iterations=30000,
            restarts=3, balance_weight=0.015,
            tmax_cap=tmax_star, relax=RELAX_TOL,
        )
        times2 = [m.total for m in m2]
        print(f"    Tmax*={tmax_star/3600:.4f}h, 容差上限={tmax_star*(1+RELAX_TOL)/3600:.4f}h, "
              f"结果 Tmax={max(times2)/3600:.4f}h, delta={(max(times2)-min(times2))/3600:.4f}h")
        q2_routes[sheet], q2_metrics[sheet] = r2, m2

        print(f"[{sheet}] 问题3：分时禁飞区事件仿真与重调度")
        zones = load_zones(sheet)
        metric3 = lambda route, c=coords, z=zones: dynamic_metric(route, c, z)
        cyclic_routes, cyclic_metrics = cyclic_improve(r2, metric3)
        cyclic_times = [m.total for m in cyclic_metrics]
        print(
            f"    循环切入优化：Tmax={max(cyclic_times) / 3600:.4f}h, "
            f"spread={(max(cyclic_times) - min(cyclic_times)) / 3600:.4f}h"
        )
        r3, m3 = optimise(
            cyclic_routes, metric3, seed=20260916 + idx, iterations=50000,
            restarts=4, balance_weight=0.025,
        )
        q3_routes[sheet], q3_metrics[sheet] = r3, m3

    write_workbook(RESULT2, q2_routes)
    write_workbook(RESULT3, q3_routes)
    os.makedirs(os.path.dirname(SUMMARY), exist_ok=True)
    with open(SUMMARY, "w", encoding="utf-8") as f:
        f.write("问题2（固定问题1机队规模；分层优化：Tmax* 容差上限内最小化 delta）\n")
        for sheet in q2_routes:
            f.write(summary_line(sheet, q2_routes[sheet], q2_metrics[sheet]))
        f.write("\n问题3（08:00出发；等待直飞与圆外切线绕飞取较早到达）\n")
        for sheet in q3_routes:
            f.write(summary_line(sheet, q3_routes[sheet], q3_metrics[sheet]))
    print(f"已写入 {RESULT2}\n已写入 {RESULT3}\n已写入 {SUMMARY}")


if __name__ == "__main__":
    main()