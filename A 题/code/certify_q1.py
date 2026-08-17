# -*- coding: utf-8 -*-
"""A 题问题一：最少无人机数量的严格下界与证书计算。

该脚本不替代 ``solve_q1.py`` 的可行路线搜索，而是为其提供下界证据：

1. ``assignment_cycle_cover_bound`` 把基地复制 N 份，求一个允许任务子环的
   最小费用循环覆盖。它忽略了“每架无人机各自连通”和 9 h 负载分配，因而
   所得总时间不大于任何真实 N 机方案的总工作时间。
2. ``solve_minmax_cycle_cover`` 进一步保留每架无人机的 9 h 工作时长约束，
   但仍允许不与基地连通的任务子环。它是原问题的严格整数松弛；若其最优
   最大工作时长大于 9 h，或者在加入 9 h 上限后被证明不可行，则真实问题
   对该 N 必然不可行。
3. 为上述模型启用 ``connected=True`` 后，增加单商品流连通约束。此时每架
   无人机的整数弧构成与基地连通的欧拉多重图，可直接还原为一条真实闭合
   路线，因而得到的是官方规则下的精确可行性模型。
4. ``solve_aggregate_total_time`` 聚合无人机编号，精确求 N 条闭合路线的
   最小总工时。若该值（或分支定界的严格下界）超过 N×9 h，即可快速排除 N。

官方补充规则通过禁止同一物理点之间的直接弧实现：同一架无人机连续停留
只算一次巡检；不同无人机访问同一点则分别计次。

依赖：numpy, openpyxl, scipy>=1.9（使用 scipy.optimize.milp/HiGHS）

示例：
    python certify_q1.py --quick
    python certify_q1.py --case Case1 --n 3 --mip --time-limit 600
    python certify_q1.py --case Case3 --n 4 --mip --time-limit 1800
    python certify_q1.py --case Case1 --n 3 --mip --exact --time-limit 1800
"""

from __future__ import annotations

import argparse
import math
import os
from dataclasses import dataclass

import numpy as np
import openpyxl
from scipy.optimize import Bounds, LinearConstraint, linear_sum_assignment, milp
from scipy.sparse import coo_matrix


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA1 = os.path.join(BASE_DIR, "附件1.xlsx")

SPEED_MS = 55.0 / 3.6
SERVICE_S = 5.0 * 60.0
SCALE_M = 100.0
TIME_LIMIT_S = 9.0 * 3600.0
LEVEL_TIMES = {"I": 3, "II": 2, "III": 1}


@dataclass(frozen=True)
class CaseData:
    name: str
    ids: np.ndarray
    coords_m: np.ndarray
    demand: np.ndarray

    @property
    def point_count(self) -> int:
        return len(self.ids)

    @property
    def visit_count(self) -> int:
        return int(self.demand.sum())


@dataclass(frozen=True)
class MipCertificate:
    case: str
    n_uav: int
    exact: bool
    status: int
    message: str
    feasible_under_9h: bool | None
    objective_h: float | None
    dual_bound_h: float | None
    mip_gap: float | None
    node_count: int | None


@dataclass(frozen=True)
class TotalTimeCertificate:
    case: str
    n_uav: int
    status: int
    message: str
    incumbent_h: float | None
    dual_bound_h: float | None
    excluded_by_bound: bool
    mip_gap: float | None
    node_count: int | None
    best_route_hours: tuple[float, ...] | None
    best_routes: tuple[tuple[int, ...], ...] | None


def load_case(sheet: str) -> CaseData:
    """读取一个算例并保留物理点层面的需求次数。"""
    wb = openpyxl.load_workbook(DATA1, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    finally:
        wb.close()
    ids = np.asarray([int(r[0]) for r in rows], dtype=int)
    coords_m = np.asarray([(float(r[1]), float(r[2])) for r in rows], dtype=float) * SCALE_M
    demand = np.asarray([LEVEL_TIMES[str(r[3]).strip()] for r in rows], dtype=int)
    return CaseData(sheet, ids, coords_m, demand)


def _distance_data(case: CaseData) -> tuple[np.ndarray, np.ndarray]:
    point_dist = np.sqrt(
        ((case.coords_m[:, None, :] - case.coords_m[None, :, :]) ** 2).sum(axis=2)
    )
    depot_dist = np.sqrt((case.coords_m**2).sum(axis=1))
    return point_dist, depot_dist


def assignment_cycle_cover_bound(case: CaseData, n_uav: int) -> tuple[float, float]:
    """返回 N 机循环覆盖的严格总距离/总工作时间下界。

    将每次巡检展开成任务副本，并把基地复制 N 份。禁止同一物理点的任务
    副本直接相连，然后求最小费用完美匹配。匹配只保留必要的入度/出度
    条件，允许分离子环，故结果是原问题总飞行距离的严格下界。
    """
    copy_point = np.repeat(np.arange(case.point_count), case.demand)
    copy_xy = case.coords_m[copy_point]
    m = len(copy_point)
    size = m + n_uav
    cost = np.full((size, size), np.inf, dtype=float)

    task_dist = np.sqrt(((copy_xy[:, None, :] - copy_xy[None, :, :]) ** 2).sum(axis=2))
    allowed = copy_point[:, None] != copy_point[None, :]
    cost[:m, :m][allowed] = task_dist[allowed]

    depot_dist = np.sqrt((copy_xy**2).sum(axis=1))
    cost[:m, m:] = depot_dist[:, None]
    cost[m:, :m] = depot_dist[None, :]
    # 基地副本之间不连边，确保每个基地副本各连接一条出弧和一条入弧。

    row_ind, col_ind = linear_sum_assignment(cost)
    distance_lb_m = float(cost[row_ind, col_ind].sum())
    total_time_lb_s = case.visit_count * SERVICE_S + distance_lb_m / SPEED_MS
    return distance_lb_m, total_time_lb_s


class _Rows:
    """稀疏线性约束矩阵构造器。"""

    def __init__(self) -> None:
        self.row: list[int] = []
        self.col: list[int] = []
        self.data: list[float] = []
        self.lb: list[float] = []
        self.ub: list[float] = []

    def add(self, terms, lb: float, ub: float) -> None:
        r = len(self.lb)
        for c, value in terms:
            if value:
                self.row.append(r)
                self.col.append(int(c))
                self.data.append(float(value))
        self.lb.append(float(lb))
        self.ub.append(float(ub))

    def constraint(self, nvars: int) -> LinearConstraint:
        matrix = coo_matrix(
            (self.data, (self.row, self.col)), shape=(len(self.lb), nvars)
        ).tocsr()
        return LinearConstraint(matrix, np.asarray(self.lb), np.asarray(self.ub))


def _best_euler_decomposition(
    case: CaseData,
    n_uav: int,
    arcs: list[tuple[int, int]],
    x_values: np.ndarray,
    *,
    trials: int = 5000,
    seed: int = 20260817,
) -> tuple[tuple[float, ...], tuple[tuple[int, ...], ...]] | None:
    """将聚合整数弧随机还原为欧拉回路，寻找最均衡的 N 段基地路线。"""
    multiplicity = np.rint(x_values).astype(int)
    edge_count = int(multiplicity.sum())
    if edge_count <= 0:
        return None

    base_adj: list[list[int]] = [[] for _ in range(case.point_count + 1)]
    for (i, j), count in zip(arcs, multiplicity):
        if count > 0:
            base_adj[i].extend([j] * int(count))

    point_dist, depot_dist = _distance_data(case)

    def route_hours(route: list[int]) -> float:
        distance_m = 0.0
        for i, j in zip(route, route[1:]):
            if i == 0:
                distance_m += depot_dist[j - 1]
            elif j == 0:
                distance_m += depot_dist[i - 1]
            else:
                distance_m += point_dist[i - 1, j - 1]
        visits = sum(node != 0 for node in route[1:])
        return visits * SERVICE_S / 3600.0 + distance_m / SPEED_MS / 3600.0

    rng = np.random.default_rng(seed)
    best_key = None
    best_hours = None
    best_routes = None
    for _ in range(max(1, trials)):
        adj = [neighbors.copy() for neighbors in base_adj]
        for neighbors in adj:
            rng.shuffle(neighbors)

        # Hierholzer 算法；随机邻接顺序给出不同但均合法的欧拉回路。
        stack = [0]
        reverse_circuit: list[int] = []
        while stack:
            node = stack[-1]
            if adj[node]:
                stack.append(adj[node].pop())
            else:
                reverse_circuit.append(stack.pop())
        circuit = reverse_circuit[::-1]
        if len(circuit) != edge_count + 1 or circuit[0] != 0 or circuit[-1] != 0:
            continue

        routes: list[list[int]] = []
        current = [0]
        for node in circuit[1:]:
            current.append(node)
            if node == 0:
                routes.append(current)
                current = [0]
        if len(routes) != n_uav:
            continue

        hours = tuple(route_hours(route) for route in routes)
        key = (max(hours), float(np.std(hours)), sum(hours))
        if best_key is None or key < best_key:
            best_key = key
            best_hours = hours
            best_routes = tuple(
                tuple(0 if node == 0 else int(case.ids[node - 1]) for node in route)
                for route in routes
            )
            if max(hours) <= 9.0 + 1e-9:
                break

    if best_hours is None or best_routes is None:
        return None
    return best_hours, best_routes


def solve_aggregate_total_time(
    case: CaseData,
    n_uav: int,
    *,
    time_limit: float = 600.0,
    mip_rel_gap: float = 0.0,
) -> TotalTimeCertificate:
    """精确求 N 条基地闭合路线的最小总工时。

    聚合整数弧 x[i,j] 的入度等于巡检需求，基地入度和出度均为 N；单商品流
    保证所有正需求节点与基地连通。平衡且连通的有向多重图存在欧拉回路，
    再在每次经过基地处切分，恰好得到 N 条不含中途返航的闭合路线。因此该
    模型精确描述 N 条路线的最小总距离，但不施加各路线 9 h 上限。

    任何满足单机 9 h 的方案总工时不超过 N×9 h。因此，只要最优值或求解器
    的全局 dual bound 大于 N×9 h，就构成候选 N 不可行的严格证书。
    """
    p = case.point_count
    nodes = range(p + 1)
    arcs = [(i, j) for i in nodes for j in nodes if i != j]
    arc_pos = {arc: a for a, arc in enumerate(arcs)}
    a_count = len(arcs)
    x_count = a_count
    f_start = x_count
    nvars = 2 * a_count

    point_dist, depot_dist = _distance_data(case)
    travel_h = np.empty(a_count, dtype=float)
    objective = np.zeros(nvars)
    lower = np.zeros(nvars)
    upper = np.zeros(nvars)
    integrality = np.zeros(nvars, dtype=np.uint8)
    integrality[:x_count] = 1

    for a, (i, j) in enumerate(arcs):
        if i == 0:
            distance = depot_dist[j - 1]
            arc_upper = min(n_uav, int(case.demand[j - 1]))
        elif j == 0:
            distance = depot_dist[i - 1]
            arc_upper = min(n_uav, int(case.demand[i - 1]))
        else:
            distance = point_dist[i - 1, j - 1]
            arc_upper = min(int(case.demand[i - 1]), int(case.demand[j - 1]))
        travel_h[a] = distance / SPEED_MS / 3600.0
        objective[a] = travel_h[a]
        upper[a] = arc_upper
        upper[f_start + a] = case.visit_count

    rows = _Rows()

    # 固定任务覆盖次数与物理点弧流守恒。
    for j in range(1, p + 1):
        incoming = [(arc_pos[(i, j)], 1.0) for i in nodes if i != j]
        demand = int(case.demand[j - 1])
        rows.add(incoming, demand, demand)
        balance = incoming + [
            (arc_pos[(j, i)], -1.0) for i in nodes if i != j
        ]
        rows.add(balance, 0.0, 0.0)

    rows.add(((arc_pos[(0, j)], 1.0) for j in range(1, p + 1)), n_uav, n_uav)
    rows.add(((arc_pos[(i, 0)], 1.0) for i in range(1, p + 1)), n_uav, n_uav)

    # 单商品流：基地向每次巡检发送一个单位，保证所有任务弧分量连接基地。
    big_m = case.visit_count
    for i, j in arcs:
        a = arc_pos[(i, j)]
        rows.add([(f_start + a, 1.0), (a, -big_m)], -np.inf, 0.0)
    for j in range(1, p + 1):
        terms = []
        for i in nodes:
            if i == j:
                continue
            terms.append((f_start + arc_pos[(i, j)], 1.0))
            terms.append((f_start + arc_pos[(j, i)], -1.0))
        demand = int(case.demand[j - 1])
        rows.add(terms, demand, demand)

    result = milp(
        c=objective,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=rows.constraint(nvars),
        options={
            "disp": True,
            "time_limit": float(time_limit),
            "mip_rel_gap": float(mip_rel_gap),
            "presolve": True,
        },
    )

    service_h = case.visit_count * SERVICE_S / 3600.0
    incumbent_h = None
    if result.fun is not None and np.isfinite(result.fun):
        incumbent_h = float(result.fun) + service_h
    raw_bound = getattr(result, "mip_dual_bound", None)
    dual_bound_h = None
    if raw_bound is not None and np.isfinite(raw_bound):
        dual_bound_h = float(raw_bound) + service_h
    # 最优结束时 scipy/HiGHS 有时不单独返回 dual bound，此时最优值本身即下界。
    if result.status == 0 and incumbent_h is not None:
        dual_bound_h = incumbent_h

    decomposition = None
    if result.x is not None:
        decomposition = _best_euler_decomposition(
            case, n_uav, arcs, np.asarray(result.x[:x_count]), trials=5000
        )
    best_route_hours = None if decomposition is None else decomposition[0]
    best_routes = None if decomposition is None else decomposition[1]

    excluded = dual_bound_h is not None and dual_bound_h > n_uav * 9.0 + 1e-8
    return TotalTimeCertificate(
        case=case.name,
        n_uav=n_uav,
        status=int(result.status),
        message=str(result.message),
        incumbent_h=incumbent_h,
        dual_bound_h=dual_bound_h,
        excluded_by_bound=excluded,
        mip_gap=None if getattr(result, "mip_gap", None) is None else float(result.mip_gap),
        node_count=None
        if getattr(result, "mip_node_count", None) is None
        else int(result.mip_node_count),
        best_route_hours=best_route_hours,
        best_routes=best_routes,
    )


def solve_minmax_cycle_cover(
    case: CaseData,
    n_uav: int,
    *,
    time_limit: float = 600.0,
    feasibility_only: bool = True,
    connected: bool = False,
    strengthened: bool = True,
    mip_rel_gap: float = 0.0,
    cap_s: float | None = None,
) -> MipCertificate:
    """求按无人机分解的整数循环覆盖松弛。

    节点 0 为基地，1..P 为物理巡检点。整数变量 x[k,i,j] 表示无人机 k
    从 i 到 j 的次数；不建立 i==j 的弧，因此连续停留不能增加巡检次数。
    模型保留每架无人机的流守恒、一次离开/返回基地、巡检总次数和工作时长，
    默认不加连通性约束，允许任务子环，因此其可行域包含所有真实路线。
    connected=True 时增加单商品流：基地发出的流量在每次任务到达处消耗 1，
    且流只能沿已使用的弧传递。这保证每个有任务的连通分量均含基地。结合
    整数弧流守恒，可得到一条经过全部弧的欧拉闭合路线，模型因而是精确的。

    feasibility_only=True 时直接加入每机 cap 上限（默认 9 h，可用 cap_s 指定）：
    若 HiGHS 返回 status=2（已证明不可行），即可严格排除该 N。否则最小化松弛问题的 Tmax。
    """
    cap = TIME_LIMIT_S if cap_s is None else float(cap_s)
    p = case.point_count
    nodes = range(p + 1)
    arcs = [(i, j) for i in nodes for j in nodes if i != j]
    arc_pos = {arc: a for a, arc in enumerate(arcs)}
    a_count = len(arcs)

    def x_index(k: int, i: int, j: int) -> int:
        return k * a_count + arc_pos[(i, j)]

    x_count = n_uav * a_count
    f_start = x_count if connected else None
    f_count = x_count if connected else 0
    # 强化连通模型时，为每架无人机显式记录巡检到达次数。q 由整数 x 唯一
    # 决定，本身可保持连续；它使径向下界只需稀疏地写入一次服务时间项。
    q_start = x_count + f_count if connected and strengthened else None
    q_count = n_uav if q_start is not None else 0
    t_index = x_count + f_count + q_count if not feasibility_only else None
    nvars = x_count + f_count + q_count + (0 if feasibility_only else 1)

    objective = np.zeros(nvars)
    if t_index is not None:
        objective[t_index] = 1.0 / 3600.0  # 目标值直接以小时表示

    lower = np.zeros(nvars)
    upper = np.zeros(nvars)
    integrality = np.zeros(nvars, dtype=np.uint8)
    integrality[:x_count] = 1
    point_dist, depot_dist = _distance_data(case)

    travel_s = np.empty(a_count, dtype=float)
    arrival_service_s = np.empty(a_count, dtype=float)
    for a, (i, j) in enumerate(arcs):
        if i == 0:
            distance = depot_dist[j - 1]
        elif j == 0:
            distance = depot_dist[i - 1]
        else:
            distance = point_dist[i - 1, j - 1]
        travel_s[a] = distance / SPEED_MS
        arrival_service_s[a] = 0.0 if j == 0 else SERVICE_S

        if i == 0 or j == 0:
            arc_upper = 1
        else:
            arc_upper = min(case.demand[i - 1], case.demand[j - 1])
        for k in range(n_uav):
            upper[k * a_count + a] = arc_upper

    if connected:
        # 9 h 内至多完成 floor(9h/5min)=108 次巡检；用它替代全局 140
        # 作为流容量，可在不改变整数可行域的前提下明显收紧 LP 松弛。
        max_route_visits = (
            int(cap // SERVICE_S) if feasibility_only else case.visit_count
        )
        upper[f_start : f_start + f_count] = max_route_visits

    if q_start is not None:
        upper[q_start : q_start + q_count] = (
            int(cap // SERVICE_S) if feasibility_only else case.visit_count
        )

    if t_index is not None:
        upper[t_index] = TIME_LIMIT_S if feasibility_only else np.inf
        integrality[t_index] = 0

    rows = _Rows()

    # 每个物理点的总进入次数等于其等级要求。
    for j in range(1, p + 1):
        terms = (
            (x_index(k, i, j), 1.0)
            for k in range(n_uav)
            for i in nodes
            if i != j
        )
        demand = int(case.demand[j - 1])
        rows.add(terms, demand, demand)

    # 每架无人机在各物理点流守恒。
    for k in range(n_uav):
        for i in range(1, p + 1):
            terms = []
            for j in nodes:
                if j == i:
                    continue
                terms.append((x_index(k, j, i), 1.0))
                terms.append((x_index(k, i, j), -1.0))
            rows.add(terms, 0.0, 0.0)

    # 固定使用 N 架无人机，每架恰好离开、返回基地一次。
    for k in range(n_uav):
        rows.add(((x_index(k, 0, j), 1.0) for j in range(1, p + 1)), 1.0, 1.0)
        rows.add(((x_index(k, i, 0), 1.0) for i in range(1, p + 1)), 1.0, 1.0)

    # 单商品流连通约束。f[k,i,j] <= M*x[k,i,j]，且每次任务到达消耗 1 单位流。
    if connected:
        def f_index(k: int, i: int, j: int) -> int:
            return f_start + k * a_count + arc_pos[(i, j)]

        big_m = int(cap // SERVICE_S) if feasibility_only else case.visit_count
        for k in range(n_uav):
            for i, j in arcs:
                rows.add(
                    [(f_index(k, i, j), 1.0), (x_index(k, i, j), -big_m)],
                    -np.inf,
                    0.0,
                )
            for j in range(1, p + 1):
                terms = []
                for i in nodes:
                    if i == j:
                        continue
                    terms.append((f_index(k, i, j), 1.0))
                    terms.append((f_index(k, j, i), -1.0))
                    terms.append((x_index(k, i, j), -1.0))
                rows.add(terms, 0.0, 0.0)

    if q_start is not None:
        # q[k] = 无人机 k 的总巡检到达次数。
        for k in range(n_uav):
            terms = [(q_start + k, 1.0)]
            terms.extend(
                (x_index(k, i, j), -1.0)
                for j in range(1, p + 1)
                for i in nodes
                if i != j
            )
            rows.add(terms, 0.0, 0.0)

        # 径向有效不等式。若 k 在点 j 承担 m 次巡检，则它至少承担 m/d_j
        # 个“访问指示量”；任何访问 j 的闭合路线飞行距离均不少于基地到 j
        # 的两倍。因此
        #   300*q_k + (2*r_j/v)*(m/d_j) <= T_k <= Tmax
        # 对所有整数可行路线均成立，并强化分数解。
        for k in range(n_uav):
            for j in range(1, p + 1):
                radial_s = 2.0 * depot_dist[j - 1] / SPEED_MS
                demand_j = float(case.demand[j - 1])
                terms = [(q_start + k, SERVICE_S)]
                terms.extend(
                    (x_index(k, i, j), radial_s / demand_j)
                    for i in nodes
                    if i != j
                )
                if feasibility_only:
                    rows.add(terms, -np.inf, cap)
                else:
                    rows.add(terms + [(t_index, -1.0)], -np.inf, 0.0)

        # 多尺度连通割。令 S 为点 h 周围最近的若干物理点。若路线访问 h，
        # 则必须至少一次离开 S；而 k 在 h 的到达次数不超过 d_h，故
        #   d_h * x_k(delta^+(S)) >= arrivals_k(h)
        # 是严格有效的。SCF 只给出 x(delta(S)) >= visits(S)/M，这组割在
        # 小邻域上强得多，且无需指数枚举全部子集。
        neighbor_order = np.argsort(point_dist, axis=1)
        # 2 点和 4 点邻域在本题上提供了大部分局部连通强化；继续加入更大
        # 邻域会把非零元从约 70 万推高到近 300 万，反而拖慢根 LP。
        neighborhood_sizes = tuple(s for s in (2, 4) if s < p)
        all_nodes = set(nodes)
        for center in range(1, p + 1):
            demand_center = float(case.demand[center - 1])
            for size in neighborhood_sizes:
                subset = {int(v) + 1 for v in neighbor_order[center - 1, :size]}
                outside = all_nodes - subset
                for k in range(n_uav):
                    terms = [
                        (x_index(k, i, j), demand_center)
                        for i in subset
                        for j in outside
                    ]
                    terms.extend(
                        (x_index(k, i, center), -1.0)
                        for i in nodes
                        if i != center
                    )
                    rows.add(terms, 0.0, np.inf)

    # 单机工作时长；服务时间按“进入一个物理点一次即巡检一次”计算。
    time_terms_by_uav: list[list[tuple[int, float]]] = []
    for k in range(n_uav):
        terms = [
            (k * a_count + a, travel_s[a] + arrival_service_s[a])
            for a in range(a_count)
        ]
        time_terms_by_uav.append(terms)
        if feasibility_only:
            rows.add(terms, -np.inf, cap)
        else:
            rows.add(terms + [(t_index, -1.0)], -np.inf, 0.0)

    # 对称性消除：按工作时长非增序排列无人机。
    for k in range(n_uav - 1):
        terms = time_terms_by_uav[k] + [
            (idx, -value) for idx, value in time_terms_by_uav[k + 1]
        ]
        rows.add(terms, 0.0, np.inf)

    options = {
        "disp": True,
        "time_limit": float(time_limit),
        "mip_rel_gap": float(mip_rel_gap),
        "presolve": True,
    }
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=rows.constraint(nvars),
        options=options,
    )

    objective_h = None
    if result.fun is not None and np.isfinite(result.fun):
        objective_h = float(result.fun) if not feasibility_only else None
    dual_bound_h = getattr(result, "mip_dual_bound", None)
    if dual_bound_h is not None and not np.isfinite(dual_bound_h):
        dual_bound_h = None

    if feasibility_only:
        if result.status == 2:
            feasible = False
        elif result.x is not None:
            feasible = True  # 仅表示松弛模型可行，不能证明真实路径可行
        else:
            feasible = None
    else:
        feasible = None

    return MipCertificate(
        case=case.name,
        n_uav=n_uav,
        exact=connected,
        status=int(result.status),
        message=str(result.message),
        feasible_under_9h=feasible,
        objective_h=objective_h,
        dual_bound_h=None if dual_bound_h is None else float(dual_bound_h),
        mip_gap=None if getattr(result, "mip_gap", None) is None else float(result.mip_gap),
        node_count=None
        if getattr(result, "mip_node_count", None) is None
        else int(result.mip_node_count),
    )


def quick_report() -> None:
    wb = openpyxl.load_workbook(DATA1, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    print("严格循环覆盖总工作量下界（> N×9 h 即可排除该 N）")
    print("case   N  visits  distance_lb(km)  total_lb(h)  avg_lb(h)  excluded")
    for sheet in sheets:
        case = load_case(sheet)
        for n_uav in range(1, 7):
            distance_m, total_s = assignment_cycle_cover_bound(case, n_uav)
            excluded = total_s > n_uav * TIME_LIMIT_S + 1e-7
            print(
                f"{sheet:<6} {n_uav:>2} {case.visit_count:>7} "
                f"{distance_m / 1000:>16.3f} {total_s / 3600:>12.4f} "
                f"{total_s / n_uav / 3600:>10.4f} {str(excluded):>9}"
            )
        print()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--quick", action="store_true", help="输出全部算例的快速严格下界")
    parser.add_argument("--case", choices=["Case1", "Case2", "Case3", "Case4"])
    parser.add_argument("--n", type=int, help="待验证的无人机数量")
    parser.add_argument("--mip", action="store_true", help="求按无人机分解的整数循环覆盖松弛")
    parser.add_argument("--aggregate", action="store_true", help="精确求 N 条路线的最小总工时")
    parser.add_argument("--exact", action="store_true", help="加入单商品流连通约束，求精确路径可行性模型")
    parser.add_argument("--optimize", action="store_true", help="最小化松弛 Tmax，而非仅验证 9 h 可行性")
    parser.add_argument("--time-limit", type=float, default=600.0, help="HiGHS 时间上限（秒）")
    parser.add_argument("--cap", type=float, default=None, help="单机工作时长上限（小时），默认 9")
    parser.add_argument("--mip-gap", type=float, default=0.0, help="允许的相对 MIP gap")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.quick or (not args.mip and not args.aggregate):
        quick_report()
    if args.aggregate:
        if args.case is None or args.n is None:
            raise SystemExit("--aggregate 必须同时指定 --case 与 --n")
        case = load_case(args.case)
        print(f"\n求解聚合连通最小总工时: {case.name}, N={args.n}, visits={case.visit_count}")
        cert = solve_aggregate_total_time(
            case, args.n, time_limit=args.time_limit, mip_rel_gap=args.mip_gap
        )
        print(cert)
        if cert.excluded_by_bound:
            print(
                f"严格证书：总工时全局下界 {cert.dual_bound_h:.6f} h "
                f"> {args.n * 9:.6f} h，因此 N={args.n} 不可行。"
            )
        else:
            print("当前总工时下界尚未超过 N×9 h，不能仅据此排除该 N。")
        if cert.best_route_hours is not None:
            print("欧拉分解路线工时(h):", tuple(round(v, 6) for v in cert.best_route_hours))
            print("欧拉分解最大工时(h):", round(max(cert.best_route_hours), 6))
            if max(cert.best_route_hours) <= 9.0 + 1e-8:
                print("严格可行证书：聚合整数弧已分解为 N 条均不超过 9 h 的真实闭合路线。")
                for idx, route in enumerate(cert.best_routes, 1):
                    print(f"UAV{idx}: " + " -> ".join(map(str, route)))
    if args.mip:
        if args.case is None or args.n is None:
            raise SystemExit("--mip 必须同时指定 --case 与 --n")
        case = load_case(args.case)
        print(
            f"\n求解整数循环覆盖松弛: {case.name}, N={args.n}, "
            f"visits={case.visit_count}, mode={'optimize' if args.optimize else '9h-feasibility'}"
        )
        cert = solve_minmax_cycle_cover(
            case,
            args.n,
            time_limit=args.time_limit,
            feasibility_only=not args.optimize,
            connected=args.exact,
            mip_rel_gap=args.mip_gap,
            cap_s=None if args.cap is None else args.cap * 3600.0,
        )
        print(cert)
        if cert.feasible_under_9h is False:
            model = "精确连通模型" if cert.exact else "循环覆盖松弛"
            print(f"严格证书：{model}已不可行，因此原路径问题在该 N 下必然不可行。")
        elif cert.feasible_under_9h is True:
            if cert.exact:
                print("精确连通模型存在 9 h 内解：该 N 在真实路径问题中可行。")
            else:
                print("松弛模型存在 9 h 内解；这既不证明也不否定原路径问题可行。")
        else:
            print("未在给定时间内形成 9 h 可行性结论，请延长时间或查看 dual bound。")


if __name__ == "__main__":
    main()
