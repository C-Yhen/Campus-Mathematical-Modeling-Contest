# -*- coding: utf-8 -*-
"""
A 题 · 问题 1 求解程序
低空经济背景下多无人机协同巡检路径优化
------------------------------------------------
功能：
  1) 计算理论下界（作业时间下界 + MST 移动下界），确定最少无人机数 Nmin 的搜索起点；
  2) 判定 Nmin：在 9 小时内完成全部巡检任务所需的最少无人机数量；
  3) 在 Nmin 架无人机下最小化总体完成时间 Tmax。主求解器为路径空间搜索
     （k-means 分组 + 路径内 2-opt + 路径间搬迁/交换 + 扰动重启，多起点运行）；
     文件中另保留序列编码遗传算法 run_ga 作为备选方案（未被主流程调用）；
  4) 输出结果表（N, Tmax, Tmin）并写入 result1.xlsx（每个算例一个 sheet）。

规则说明（与论文假设一致）：
  - 同一架无人机路径中相邻重复访问同一原始任务点，只计 1 次作业时间（连续停留算 1 次巡检）；
  - 不同无人机（或同一架离开后再次返回）访问同一任务点，各计 1 次巡检。

运行：python solve_q1.py
依赖：numpy, openpyxl
"""

import math
import os
import random
import time
from collections import Counter

import numpy as np
import openpyxl

# ---------------------------- 基本常量 ----------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA1 = os.path.join(BASE_DIR, "附件1.xlsx")
RESULT1 = os.path.join(BASE_DIR, "result1.xlsx")

SPEED_KMH = 55.0                    # 平均巡航速度 (km/h)
SPEED_MS = SPEED_KMH / 3.6          # (m/s)
SERVICE_S = 5.0 * 60.0              # 单点单次作业时间 (s)
SCALE_M = 100.0                     # 坐标单位换算: 1 单位 = 100 m
TIME_LIMIT_S = 9.0 * 3600.0         # 9 小时硬约束 (s)
LEVEL_TIMES = {"I": 3, "II": 2, "III": 1}   # 巡检等级 -> 所需次数

# ---------------------------- 数据读取 ----------------------------
def load_case(sheet):
    """读取附件1 中某个算例: 返回 (原始点列表, 副本列表)
    原始点: (Point_ID, x, y, level)
    副本:   每个原始点按等级复制 k 份 -> [(orig_id, x, y), ...]
    """
    wb = openpyxl.load_workbook(DATA1)
    ws = wb[sheet]
    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        pid, x, y, lvl = int(row[0]), float(row[1]), float(row[2]), str(row[3]).strip()
        points.append((pid, x, y, lvl))
    wb.close()
    copies = []
    for pid, x, y, lvl in points:
        for _ in range(LEVEL_TIMES.get(lvl, 1)):
            copies.append((pid, x, y))
    return points, copies


def build_dist(copies):
    """距离矩阵 (m)。索引 0 表示基地 (0,0)，索引 1..M 对应副本点。"""
    coords = np.array([[0.0, 0.0]] + [[c[1], c[2]] for c in copies], dtype=float)
    d = np.sqrt(((coords[:, None, :] - coords[None, :, :]) ** 2).sum(-1)) * SCALE_M
    return d


# ---------------------------- 时间计算 ----------------------------
def route_seconds(route, dist, orig_of):
    """单条路径总耗时 (s)。route 为副本索引列表，首尾默认基地。
    规则：相邻两个副本若属于同一原始点，则只计 1 次作业时间。
    """
    if not route:
        return 0.0
    t = dist[0, route[0]] / SPEED_MS + SERVICE_S
    for a, b in zip(route, route[1:]):
        t += dist[a, b] / SPEED_MS
        if orig_of[b] != orig_of[a]:
            t += SERVICE_S
    t += dist[route[-1], 0] / SPEED_MS
    return t


def full_time(seq, dist, orig_of):
    """单机串行访问 seq 全部点的总耗时 (s)，用于负载切分基准。"""
    return route_seconds(seq, dist, orig_of)


def split_tour(seq, dist, orig_of, N):
    """把巨型序列 seq 按时间负载贪心切成 N 条路径（每条从基地出发返回基地）。"""
    target = full_time(seq, dist, orig_of) / N
    routes, cur = [], []
    cur_open = 0.0        # 当前开放路径耗时（基地->...->最后点，含作业）
    last = 0
    last_pid = None
    for i in seq:
        move = dist[last, i] / SPEED_MS
        serv = SERVICE_S if orig_of[i] != last_pid else 0.0
        close = dist[i, 0] / SPEED_MS
        if cur and len(routes) < N - 1 and cur_open + move + serv + close > target:
            routes.append(cur)                     # 切分
            cur = [i]
            cur_open = dist[0, i] / SPEED_MS + SERVICE_S
            last_pid = orig_of[i]
        else:
            cur.append(i)
            cur_open += move + serv
            last_pid = orig_of[i]
        last = i
    if cur:
        routes.append(cur)
    # 段数不足 N 时，把最长段一分为二
    while len(routes) < N:
        j = max(range(len(routes)), key=lambda k: route_seconds(routes[k], dist, orig_of))
        r = routes[j]
        m = len(r) // 2
        if m == 0:
            break                                   # 无法再分
        routes[j:j + 1] = [r[:m], r[m:]]
    return routes


def check_cover(routes, orig_of, req):
    """检查每个原始任务点的巡检次数是否满足等级要求。
    规则: 同一路径中相邻重复访问同一原始点合并计 1 次。
    """
    cnt = Counter()
    for r in routes:
        prev = None
        for i in r:
            pid = orig_of[i]
            if pid != prev:
                cnt[pid] += 1
            prev = pid
    for pid, k in req.items():
        if cnt.get(pid, 0) < k:
            return False
    return True


PENALTY = 1e7


def evaluate(seq, dist, orig_of, N, req=None):
    """评估一个巨型序列: 返回 (Tmax(秒), Tmin(秒), routes)。
    若 req 给定，则按缺失的巡检次数乘惩罚系数计入 Tmax。
    """
    routes = split_tour(seq, dist, orig_of, N)
    ts = [route_seconds(r, dist, orig_of) for r in routes]
    tmax, tmin = max(ts), min(ts)
    if req is not None:
        cnt = Counter()
        for r in routes:
            prev = None
            for i in r:
                pid = orig_of[i]
                if pid != prev:
                    cnt[pid] += 1
                prev = pid
        missing = sum(max(0, k - cnt.get(pid, 0)) for pid, k in req.items())
        tmax += missing * PENALTY
    return tmax, tmin, routes


def repair(seq, orig_of):
    """消除序列中相邻的同一原始点副本（相邻连续停留合并计 1 次 -> 相邻重复非法）。
    两步法: 先按顺序收集不同点，冲突副本暂存；再把暂存副本插入合法位置。
    """
    pending = []
    out = []
    last = None
    for i in seq:
        pid = orig_of[i]
        if pid != last:
            out.append(i)
            last = pid
        else:
            pending.append(i)
    for it in pending:
        pid = orig_of[it]
        placed = False
        for pos in range(len(out), -1, -1):
            left = orig_of[out[pos - 1]] if pos > 0 else None
            right = orig_of[out[pos]] if pos < len(out) else None
            if left != pid and right != pid:
                out.insert(pos, it)
                placed = True
                break
        if not placed:
            out.append(it)
    return out


# ---------------------------- 理论下界 ----------------------------
def mst_length(dist):
    """Prim 算法求基地+所有副本点的最小生成树总长 (m)。"""
    n = dist.shape[0]
    used = np.zeros(n, dtype=bool)
    dmin = np.full(n, np.inf)
    dmin[0] = 0.0
    total = 0.0
    for _ in range(n):
        u = int(np.argmin(np.where(used, np.inf, dmin)))
        used[u] = True
        total += dmin[u]
        cand = dist[u]
        mask = (~used) & (cand < dmin)
        dmin[mask] = cand[mask]
    return total


def lower_bound(copies, dist):
    """返回 (Nmin 搜索起点, 各下界分量)。
    LB1 = ceil(总作业时间 / 9h)
    LB2 = ceil((总作业时间 + MST移动时间) / 9h)
    """
    total_serv = len(copies) * SERVICE_S
    lb1 = math.ceil((total_serv / 3600.0) / 9.0)
    mst_t = mst_length(dist) / SPEED_MS
    lb2 = math.ceil(((total_serv + mst_t) / 3600.0) / 9.0)
    return max(lb1, lb2), (lb1, lb2, mst_t / 3600.0)


# ---------------------------- 初始解构造 ----------------------------
def sector_seq(copies, N):
    """扇区分割: 按极角排序后连续切成 N 段，返回巨型序列。
    copies: [(orig_id, x, y), ...]
    """
    M = len(copies)
    ang = [math.atan2(c[2], c[1]) for c in copies]
    order = sorted(range(1, M + 1), key=lambda i: ang[i - 1])
    base, rem = M // N, M % N
    seq, s = [], 0
    for k in range(N):
        L = base + (1 if k < rem else 0)
        seq.extend(order[s:s + L])
        s += L
    return seq


def perturb(seq, rng, k=6):
    """轻微扰动: 随机交换 k 次，用于增加初始解多样性。"""
    s = seq[:]
    for _ in range(k):
        a, b = rng.sample(range(len(s)), 2)
        s[a], s[b] = s[b], s[a]
    return s


def kmeans_seq(copies, N, rng, iters=25):
    """简单 k-means 聚类分簇，簇内按最近邻排序，簇间拼接。"""
    M = len(copies)
    coords = np.array([[c[1], c[2]] for c in copies])
    idx = list(range(1, M + 1))
    # 最远点法初始化簇心
    centers = [coords[rng.randrange(M)]]
    for _ in range(N - 1):
        dmin = np.min([np.sum((coords - c) ** 2, axis=1) for c in centers], axis=0)
        centers.append(coords[int(np.argmax(dmin))])
    centers = np.array(centers)
    labels = np.zeros(M, dtype=int)
    for _ in range(iters):
        labels = np.argmin(np.sum((coords[:, None, :] - centers[None, :, :]) ** 2, axis=2), axis=1)
        new_c = np.array([coords[labels == k].mean(axis=0) if (labels == k).any()
                          else coords[rng.randrange(M)] for k in range(N)])
        if np.allclose(new_c, centers):
            centers = new_c
            break
        centers = new_c
    seq = []
    for k in range(N):
        cluster = [i for i, lb in zip(idx, labels) if lb == k]
        # 簇内最近邻
        if cluster:
            start = min(cluster, key=lambda i: dist_base(coords[i - 1]))
            cur = start
            remain = set(cluster) - {start}
            sub = [start]
            while remain:
                nxt = min(remain, key=lambda i: math.hypot(coords[i - 1][0] - coords[cur - 1][0],
                                                           coords[i - 1][1] - coords[cur - 1][1]))
                sub.append(nxt)
                remain.remove(nxt)
                cur = nxt
            seq.extend(sub)
    return seq


def dist_base(c):
    return math.hypot(c[0], c[1])


# ---------------------------- 遗传算法 ----------------------------
def ox_crossover(a, b, rng):
    """顺序交叉 (OX)，支持重复元素。"""
    M = len(a)
    i, j = sorted(rng.sample(range(M), 2))
    mid = a[i:j]
    cnt = Counter(mid)
    rest = []
    for x in b:
        if cnt[x] > 0:
            cnt[x] -= 1
        else:
            rest.append(x)
    return rest[:i] + mid + rest[i:]


def preseq_improve(seq, dist, orig_of, rng, tries=80):
    """以单机串行总时间为目标做随机 2-opt，提升初始解质量。"""
    cur = seq[:]
    cur_t = full_time(cur, dist, orig_of)
    for _ in range(tries):
        a, b = sorted(rng.sample(range(len(cur)), 2))
        cand = cur[:a] + list(reversed(cur[a:b + 1])) + cur[b + 1:]
        t = full_time(cand, dist, orig_of)
        if t < cur_t - 1e-9:
            cur, cur_t = cand, t
    return cur


def mutate(seq, rng):
    """随机变异: swap / 2-opt 反转 / insert / 块搬移。"""
    s = seq[:]
    op = rng.randrange(4)
    M = len(s)
    if op == 0:
        a, b = rng.sample(range(M), 2)
        s[a], s[b] = s[b], s[a]
    elif op == 1:
        a, b = sorted(rng.sample(range(M), 2))
        s[a:b + 1] = reversed(s[a:b + 1])
    elif op == 2:
        a = rng.randrange(M)
        x = s.pop(a)
        b = rng.randrange(len(s) + 1)
        s.insert(b, x)
    else:
        a, b = sorted(rng.sample(range(M + 1), 2))
        L = min(b - a, 8)
        if L > 0:
            blk = s[a:a + L]
            rest = s[:a] + s[a + L:]
            c = rng.randrange(len(rest) + 1)
            s = rest[:c] + blk + rest[c:]
    return s


def local_search(seq, dist, orig_of, N, req, tries=120):
    """对巨型序列做 first-improve 局部搜索: 2-opt / swap / insert。"""
    M = len(seq)
    cur = seq[:]
    cur_t, _, _ = evaluate(cur, dist, orig_of, N, req)
    for _ in range(tries):
        op = random.randrange(3)
        improved = False
        for _ in range(80):
            if op == 0:      # 2-opt
                a, b = sorted(random.sample(range(M), 2))
                cand = cur[:a] + list(reversed(cur[a:b + 1])) + cur[b + 1:]
            elif op == 1:    # swap
                a, b = random.sample(range(M), 2)
                cand = cur[:]
                cand[a], cand[b] = cand[b], cand[a]
            else:            # insert
                a = random.randrange(M)
                cand = cur[:]
                x = cand.pop(a)
                b = random.randrange(len(cand) + 1)
                cand.insert(b, x)
            nt, _, _ = evaluate(cand, dist, orig_of, N, req)
            if nt < cur_t - 1e-9:
                cur, cur_t = cand, nt
                improved = True
                break
        if not improved:
            break
    return cur, cur_t


def relocate_ls(routes, dist, orig_of, iters=25, allow_swap=True):
    """路径间搬迁/交换: 把最忙路径的点搬到最闲路径（或交换两点），降低 Tmax。"""
    ts = [route_seconds(r, dist, orig_of) for r in routes]
    for _ in range(iters):
        jmax = int(np.argmax(ts))
        jmin = int(np.argmin(ts))
        if jmax == jmin or not routes[jmax]:
            break
        best_move = None
        best_t = max(ts[jmax], ts[jmin])
        r_from, r_to = routes[jmax], routes[jmin]
        for i in range(len(r_from)):
            p = r_from[i]
            pid_p = orig_of[p]
            # 移除后 from 两端不能变成相邻重复
            if i > 0 and i + 1 < len(r_from) and orig_of[r_from[i - 1]] == orig_of[r_from[i + 1]]:
                continue
            new_from = r_from[:i] + r_from[i + 1:]
            tf = route_seconds(new_from, dist, orig_of)
            if tf >= best_t:
                continue
            # 搬点: p 插入 r_to 的最佳位置
            for pos in range(len(r_to) + 1):
                if pos > 0 and orig_of[r_to[pos - 1]] == pid_p:
                    continue
                if pos < len(r_to) and orig_of[r_to[pos]] == pid_p:
                    continue
                new_to = r_to[:pos] + [p] + r_to[pos:]
                tt = route_seconds(new_to, dist, orig_of)
                if max(tf, tt) < best_t - 1e-9:
                    best_t = max(tf, tt)
                    best_move = (i, pos, None)
            # 交换: p 与 r_to 中某点 q 互换
            if allow_swap:
                for j in range(len(r_to)):
                    q = r_to[j]
                    pid_q = orig_of[q]
                    new_from2 = r_from[:i] + [q] + r_from[i + 1:]
                    new_to2 = r_to[:j] + [p] + r_to[j + 1:]
                    if (i > 0 and orig_of[r_from[i - 1]] == pid_q) or \
                       (i + 1 < len(r_from) and orig_of[r_from[i + 1]] == pid_q):
                        continue
                    if (j > 0 and orig_of[r_to[j - 1]] == pid_p) or \
                       (j + 1 < len(r_to) and orig_of[r_to[j + 1]] == pid_p):
                        continue
                    tf2 = route_seconds(new_from2, dist, orig_of)
                    tt2 = route_seconds(new_to2, dist, orig_of)
                    if max(tf2, tt2) < best_t - 1e-9:
                        best_t = max(tf2, tt2)
                        best_move = (i, -1, j)
        if best_move is None:
            break
        i, pos, jsw = best_move
        if jsw is None:
            p = routes[jmax].pop(i)
            routes[jmin].insert(pos, p)
        else:
            routes[jmax][i], routes[jmin][jsw] = routes[jmin][jsw], routes[jmax][i]
        ts = [route_seconds(r, dist, orig_of) for r in routes]
    return routes, max(ts)


def run_ga(N, dist, orig_of, copies, req, pop=80, gens=300, seed=42, time_cap=None, init_seqs=None):
    """[备选求解器，未被主流程调用] 序列编码遗传算法: 最小化 Tmax。
    主求解器为 solve_vrp_routes（路径空间搜索），本函数保留作为对比/备选。
    orig_of: 长度 M+1 的原始点编号映射（0=基地）
    copies : 长度 M 的 [(orig_id, x, y), ...]（初始解构造用）
    req    : {orig_id: 要求次数}，用于巡检次数校验
    返回 (best_Tmax 秒, best_routes)。
    """
    rng = random.Random(seed)
    M = len(orig_of) - 1
    pop_seqs = list(init_seqs or [])
    # 混合初始解: 扇区分割 / k-means 聚类 / 随机，全部 repair + 预优化
    if not pop_seqs:
        for _ in range(max(1, pop // 3)):
            s = repair(perturb(sector_seq(copies, N), rng), orig_of)
            pop_seqs.append(preseq_improve(s, dist, orig_of, rng))
        for _ in range(max(1, pop // 3)):
            s = repair(kmeans_seq(copies, N, rng), orig_of)
            pop_seqs.append(preseq_improve(s, dist, orig_of, rng))
    while len(pop_seqs) < pop:
        s = repair(rng.sample(range(1, M + 1), M), orig_of)
        pop_seqs.append(preseq_improve(s, dist, orig_of, rng))
    pop_seqs = pop_seqs[:pop]

    fits = [evaluate(s, dist, orig_of, N, req)[0] for s in pop_seqs]
    best_i = int(np.argmin(fits))
    best_t = fits[best_i]
    best_seq = pop_seqs[best_i][:]

    elite_n = max(2, pop // 10)
    stall = 0
    for gen in range(gens):
        order = sorted(range(pop), key=lambda i: fits[i])
        new = [pop_seqs[i][:] for i in order[:elite_n]]
        while len(new) < pop:
            # 锦标赛选择
            cands = [rng.randrange(pop) for _ in range(3)]
            p1 = pop_seqs[min(cands, key=lambda i: fits[i])]
            cands = [rng.randrange(pop) for _ in range(3)]
            p2 = pop_seqs[min(cands, key=lambda i: fits[i])]
            if rng.random() < 0.65:
                child = repair(ox_crossover(p1, p2, rng), orig_of)
            else:
                child = repair(mutate(p1, rng), orig_of)
            new.append(child)
        pop_seqs = new
        fits = [evaluate(s, dist, orig_of, N, req)[0] for s in pop_seqs]

        # 每代对最优 3 个个体做局部搜索
        top_idx = sorted(range(pop), key=lambda i: fits[i])[:3]
        for bi in top_idx:
            ls_seq, ls_t = local_search(pop_seqs[bi], dist, orig_of, N, req, tries=8)
            if ls_t < fits[bi] - 1e-9:
                pop_seqs[bi] = ls_seq
                fits[bi] = ls_t

        # 每 20 代对当前最优解做一次路径间搬迁平衡
        bi = int(np.argmin(fits))
        if gen % 20 == 19:
            _, _, br = evaluate(pop_seqs[bi], dist, orig_of, N, req)
            br, bt = relocate_ls(br, dist, orig_of, iters=150)
            if bt < fits[bi] - 1e-9:
                seq_flat = [i for r in br for i in r]
                pop_seqs[bi] = seq_flat
                # 注意: 压平后的序列重新切分可能改变路径组合，须重新评估
                fits[bi] = evaluate(seq_flat, dist, orig_of, N, req)[0]
        bi = int(np.argmin(fits))
        if fits[bi] < best_t - 1e-9:
            best_t = fits[bi]
            best_seq = pop_seqs[bi][:]
            stall = 0
        else:
            stall += 1

        if time_cap is not None and best_t <= time_cap:
            break
        if stall >= 120:
            break
        if gen % 50 == 0 or gen == gens - 1:
            print(f"    gen {gen + 1:4d}/{gens}  best Tmax = {best_t / 3600.0:.3f} h")

    # 最终解: 路径层搬迁改进
    _, _, routes = evaluate(best_seq, dist, orig_of, N, req)
    routes, best_t2 = relocate_ls(routes, dist, orig_of, iters=300)
    if best_t2 < best_t - 1e-9:
        best_t = best_t2
    else:
        _, _, routes = evaluate(best_seq, dist, orig_of, N, req)
    return best_t, routes


# ---------------------------- 路径空间求解器 ----------------------------
def two_opt_route_fast(route, dist, orig_of, rounds=30):
    """单路径 2-opt（first-improve，移动时间增量 O(1)）。
    反转段可延伸到路径两端（端点锚点取基地），跳过会制造相邻同点重复的反转。
    """
    r = route[:]
    n = len(r)
    if n < 4:
        return r
    for _ in range(rounds):
        improved = False
        for a in range(0, n - 1):
            pa = r[a - 1] if a > 0 else 0          # 左锚点（基地或前驱）
            for b in range(a + 1, n):
                nb = r[b + 1] if b < n - 1 else 0  # 右锚点（基地或后继）
                if orig_of[r[b]] == orig_of[pa] or orig_of[r[a]] == orig_of[nb]:
                    continue
                d = dist[pa, r[b]] + dist[r[a], nb] - dist[pa, r[a]] - dist[r[b], nb]
                if d < -1e-9:
                    r[a:b + 1] = list(reversed(r[a:b + 1]))
                    improved = True
        if not improved:
            break
    return r


def perturb_routes(routes, orig_of, rng, k=2):
    """随机扰动: 跨路径交换 k 个点，帮助跳出局部最优。"""
    for _ in range(k):
        r1, r2 = rng.randrange(len(routes)), rng.randrange(len(routes))
        if not routes[r1] or not routes[r2]:
            continue
        a = rng.randrange(len(routes[r1]))
        b = rng.randrange(len(routes[r2]))
        routes[r1][a], routes[r2][b] = routes[r2][b], routes[r1][a]
        # 交换可能造成相邻同点副本，立即修复
        routes[r1] = route_repair(routes[r1], orig_of)
        routes[r2] = route_repair(routes[r2], orig_of)
    return routes


def route_repair(route, orig_of):
    """把路径中相邻的同一原始点副本移开到路径内其他合法位置。"""
    out, pending = [], []
    last = None
    for i in route:
        pid = orig_of[i]
        if pid != last:
            out.append(i)
            last = pid
        else:
            pending.append(i)
    for it in pending:
        pid = orig_of[it]
        placed = False
        for pos in range(len(out), -1, -1):
            left = orig_of[out[pos - 1]] if pos > 0 else None
            right = orig_of[out[pos]] if pos < len(out) else None
            if left != pid and right != pid:
                out.insert(pos, it)
                placed = True
                break
        if not placed:
            out.append(it)
    return out


def solve_vrp_routes(N, dist, orig_of, req, copies, seed=42, restarts=6,
                     outer_iters=40, time_cap=None):
    """路径空间求解: k-means 分组 -> 2-opt -> 搬迁平衡 -> 扰动重启。
    返回 (best_Tmax 秒, best_routes)，若找到 <= time_cap 的解则提前停止。
    """
    best_t = float('inf')
    best_routes = None
    M = len(orig_of) - 1
    for trial in range(restarts):
        rng = random.Random(seed * 131 + trial)
        # 多样初始解: 热启动 / 扇区 / k-means / 随机
        if trial > 0 and trial % 4 == 0 and best_routes is not None:
            routes = [r[:] for r in best_routes]
        elif trial % 3 == 0:
            seq = repair(sector_seq(copies, N), orig_of)
            routes = split_tour(seq, dist, orig_of, N)
        elif trial % 3 == 1:
            seq = repair(kmeans_seq(copies, N, rng), orig_of)
            routes = split_tour(seq, dist, orig_of, N)
        else:
            seq = repair(rng.sample(range(1, M + 1), M), orig_of)
            routes = split_tour(seq, dist, orig_of, N)
        routes = [two_opt_route_fast(r, dist, orig_of, rounds=10) for r in routes]
        for it in range(outer_iters):
            routes, _ = relocate_ls(routes, dist, orig_of, iters=30)
            routes = [two_opt_route_fast(r, dist, orig_of, rounds=5) for r in routes]
            if it % 15 == 14:
                routes = perturb_routes(routes, orig_of, rng, k=2)
        # 最终精修
        routes, _ = relocate_ls(routes, dist, orig_of, iters=250)
        routes = [two_opt_route_fast(r, dist, orig_of, rounds=60) for r in routes]
        routes, t = relocate_ls(routes, dist, orig_of, iters=150)
        ts = [route_seconds(r, dist, orig_of) for r in routes]
        tmax = max(ts)
        if not check_cover(routes, orig_of, req):
            continue
        if tmax < best_t:
            best_t, best_routes = tmax, [r[:] for r in routes]
        if time_cap is not None and best_t <= time_cap:
            break
    if best_routes is None:
        raise RuntimeError(f"N={N} 下所有 trial 均未找到满足巡检次数要求的合法解")
    return best_t, best_routes


# ---------------------------- 单算例求解 ----------------------------
def solve_case(sheet, seed=42):
    print("=" * 64)
    print(f"算例 {sheet}")
    pts, copies = load_case(sheet)
    dist = build_dist(copies)
    # 长度 M+1 的原始点编号映射: 索引与 dist 对齐 (0=基地, 1..M=副本)
    orig_of = [None] + [c[0] for c in copies]
    M = len(copies)
    # 各原始点的巡检次数要求
    req = {pid: LEVEL_TIMES.get(lvl, 1) for pid, x, y, lvl in pts}

    lb, (lb1, lb2, mst_h) = lower_bound(copies, dist)
    total_serv_h = M * SERVICE_S / 3600.0
    print(f"  点数 {len(pts)} | 副本(总访问次数) {M} | 总作业 {total_serv_h:.2f} h | "
          f"MST移动 {mst_h:.2f} h")
    print(f"  下界: LB1(作业)={lb1}  LB2(作业+移动)={lb2}  -> N 从 {lb} 开始判定")

    # ---------- 判定 Nmin ----------
    Nmin = None
    best_quick = None
    margin = TIME_LIMIT_S * 1.2          # 快速判定置信阈值 (10.8h)
    for N in range(lb, lb + 4):
        print(f"  [判定] 尝试 N = {N}")
        t0 = time.time()
        best_t = float('inf')
        routes_best = None
        # 阶段1: 快速初判（多种子，小参数）
        for sd in (seed, seed + 7):
            bt, br = solve_vrp_routes(N, dist, orig_of, req, copies, seed=sd,
                                      restarts=4, outer_iters=30,
                                      time_cap=TIME_LIMIT_S)
            if bt < best_t:
                best_t, routes_best = bt, br
            if best_t <= TIME_LIMIT_S + 1e-6:
                break
        # 阶段2: 临界区间 (9h, 10.8h] 加强搜索，避免“可行但没搜到”的误判
        if TIME_LIMIT_S < best_t <= margin:
            print(f"    快速判定 {best_t / 3600.0:.3f} h 接近边界，加强搜索 ...")
            for sd in (seed + 13, seed + 19, 7):
                bt, br = solve_vrp_routes(N, dist, orig_of, req, copies, seed=sd,
                                          restarts=18, outer_iters=70,
                                          time_cap=TIME_LIMIT_S)
                if bt < best_t:
                    best_t, routes_best = bt, br
                if best_t <= TIME_LIMIT_S + 1e-6:
                    break
        print(f"    N={N}: Tmax = {best_t / 3600.0:.3f} h   (耗时 {time.time() - t0:.1f}s)")
        if best_t <= TIME_LIMIT_S + 1e-6:
            Nmin = N
            best_quick = (best_t, routes_best)
            break
    assert Nmin is not None, "未找到可行架数，请增大搜索范围"

    # ---------- 在 Nmin 下优化 Tmax ----------
    print(f"  [优化] Nmin = {Nmin}，进一步压缩 Tmax ...")
    t0 = time.time()
    # 按单机路径规模自适应: 路径长时缩减搜索量
    avg_len = M / Nmin
    if avg_len > 45:
        opt_restarts, opt_outer = 6, 30
    else:
        opt_restarts, opt_outer = 15, 70
    best_t = float('inf')
    routes = None
    for sd in (seed + 1, seed + 11, seed + 21):
        bt, br = solve_vrp_routes(Nmin, dist, orig_of, req, copies, seed=sd,
                                  restarts=opt_restarts, outer_iters=opt_outer)
        if bt < best_t:
            best_t, routes = bt, br
    # 若判定阶段结果更好则保留
    if best_quick is not None and best_quick[0] < best_t:
        best_t, routes = best_quick
    ts = [route_seconds(r, dist, orig_of) for r in routes]
    tmax, tmin = max(ts), min(ts)
    print(f"    最终: N = {Nmin}, Tmax = {tmax / 3600.0:.4f} h, "
          f"Tmin = {tmin / 3600.0:.4f} h   (耗时 {time.time() - t0:.1f}s)")
    # 将副本索引映射回原始点编号，供结果表输出
    routes_id = [[orig_of[i] for i in r] for r in routes]
    return Nmin, tmax, tmin, routes_id, pts


# ---------------------------- 输出 ----------------------------
def write_result_xlsx(results, max_len=250):
    """按模板格式写 result1.xlsx: 每个算例一个 sheet，每行一架无人机。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet, (N, tmax, tmin, routes, pts) in results.items():
        ws = wb.create_sheet(sheet)
        maxr = max((len(r) for r in routes), default=0)
        maxr = min(maxr, max_len)
        header = ["UAV ID"] + [f"{i}th Inspection Point" for i in range(1, maxr + 1)]
        ws.append(header)
        for uav, r in enumerate(routes, start=1):
            row = [uav] + [pid for pid in r[:maxr]]
            ws.append(row)
    wb.save(RESULT1)
    print(f"调度方案已写入 {RESULT1}")


def print_result_table(results):
    print("\n" + "=" * 64)
    print("问题 1 结果总表（表 2 格式）")
    print("-" * 64)
    print(f"{'测试算例':<12}{'无人机数量N':<12}{'Tmax(h)':<14}{'Tmin(h)':<12}")
    print("-" * 64)
    for sheet, (N, tmax, tmin, routes, pts) in results.items():
        print(f"{sheet:<12}{N:<12}{tmax / 3600.0:<14.4f}{tmin / 3600.0:<12.4f}")
    print("=" * 64)


def main():
    seed = 20260816
    results = {}
    for sheet in ["Case1", "Case2", "Case3", "Case4"]:
        N, tmax, tmin, routes, pts = solve_case(sheet, seed=seed)
        results[sheet] = (N, tmax, tmin, routes, pts)
    write_result_xlsx(results)
    print_result_table(results)


if __name__ == "__main__":
    main()
